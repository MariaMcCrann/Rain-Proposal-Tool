"""
Takes the extracted RFQ JSON (from extract_rfq.py) and:

1. Writes deliverables into the Rain Fee task description column, starting
   at row 12. If there are more than 10 deliverables, the template grows -
   rows are inserted before the Sub Total row, copying the same formula
   pattern, and the Sub Total/Total formulas are extended to cover the new
   rows. Hours/fee columns are left untouched either way - they stay blank
   for a human to fill in during review.
2. Fixes the Jessica Ward role mislabel in the Variation 1 block (G33),
   which said "Project Engineer" but should match her real title
   "Hydrologist and Technician" (per G10 in the main table). Nothing else
   in the Variation 1 block is touched.
3. Computes a ROUGH order-of-magnitude fee estimate as plain text, kept
   completely separate from the spreadsheet. The spreadsheet only ever
   gets task descriptions - no numbers.

Existing cell formatting (font, color, bold) is preserved automatically
because we only set .value, never .font - openpyxl doesn't reset style
unless you explicitly tell it to. For newly inserted rows, formatting is
copied explicitly from the last existing task row.
"""

import re
import copy
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

TASK_ROW_START = 12
TASK_ROW_END = 21  # rows available before any insertion is needed
DESCRIPTION_COL = "B"

# Hourly rate card straight from the template header row (C8, E8, G8, ...)
STAFF_RATES = [400, 400, 400, 320, 320, 250, 220, 180, 180, 200]

# Columns in the fee table, left to right: Item, Description, then 10 pairs
# of (Hours, Fee) per staff member, then Expenses, then row Sub Total.
HOURS_COLS = ["C", "E", "G", "I", "K", "M", "O", "Q", "S", "U"]
FEE_COLS = ["D", "F", "H", "J", "L", "N", "P", "R", "T", "V"]
EXPENSE_COL = "W"
ROW_SUBTOTAL_COL = "X"


def _shift_merged_ranges_below(ws, insertion_row: int, amount: int) -> None:
    """openpyxl's insert_rows moves cell values/styles but NOT merged cell
    ranges - this does that part manually.

    Deliberately manipulates ws.merged_cells.ranges directly rather than
    calling unmerge_cells()/merge_cells(): after insert_rows, those helper
    methods try to clean up internal cell bookkeeping that's already gone
    stale (the cells they expect at the old coordinates no longer match
    what insert_rows actually did), which throws a KeyError. Working on the
    ranges collection directly avoids touching that stale bookkeeping."""
    from openpyxl.worksheet.cell_range import CellRange

    to_shift = [r for r in list(ws.merged_cells.ranges) if r.min_row >= insertion_row]
    for r in to_shift:
        ws.merged_cells.ranges.remove(r)
    for r in to_shift:
        ws.merged_cells.ranges.add(
            CellRange(min_col=r.min_col, min_row=r.min_row + amount, max_col=r.max_col, max_row=r.max_row + amount)
        )


def _reindex_formula_self_refs(formula, old_row: int, new_row: int):
    """After insert_rows physically moves a row, formulas inside it still
    textually reference their OLD row number for same-row cells (e.g. a fee
    formula sitting at its new row 40 might still say '=C35*C$8' instead of
    '=C40*C$8'). This rewrites those relative self-references. Absolute
    references like C$8 are left untouched (the $ blocks the match)."""
    if not isinstance(formula, str) or not formula.startswith("="):
        return formula
    pattern = re.compile(r'(?<!\$)([A-Z]{1,2})' + str(old_row) + r'\b')
    return pattern.sub(lambda m: f"{m.group(1)}{new_row}", formula)


def fill_phased_template(template_path: str, output_path: str, extracted: dict) -> list[str]:
    """
    Like fill_task_descriptions, but groups tasks under phase header rows
    with a per-phase subtotal, matching how real Rain proposals are
    structured (Phase 3.1, Phase 3.2, ...). No dollar figures are written
    anywhere - hours, fees, and the final total all stay as live formulas
    evaluating to zero until a human fills in hours.

    The original template has no concept of phase headers, so this treats
    rows 12 onward as a blank canvas: it makes enough room (reusing the
    same insert/shift/reindex machinery as fill_task_descriptions), then
    writes every row in that block fresh - phase header, task rows, phase
    subtotal, repeat, then one final Total row that sums the phase
    subtotals (not a cell range, since phases aren't contiguous after the
    header rows are mixed in).
    """
    wb = load_workbook(template_path)
    ws = wb["Rain Fee"]

    phases = extracted.get("phases", [])
    warnings = []
    if not phases:
        warnings.append("No phases/deliverables extracted - task table left as the blank template default.")
        wb.save(output_path)
        return warnings

    total_needed = sum(2 + len(p.get("deliverables", [])) for p in phases) + 1  # +1 for the final Total row
    old_block_size = (TASK_ROW_END + 2) - TASK_ROW_START + 1  # rows 12-23 = 12

    # Capture style templates from the original flat block before touching
    # anything - these become the look for every row type in the new layout.
    task_style = {c.column_letter: copy.copy(ws[f"{c.column_letter}{TASK_ROW_END}"]) for c in ws[TASK_ROW_END]}
    subtotal_style = {c.column_letter: copy.copy(ws[f"{c.column_letter}{TASK_ROW_END + 1}"]) for c in ws[TASK_ROW_END + 1]}
    header_style = copy.copy(ws["B30"])  # "Variation 1" label - the template's existing section-header convention

    extra = max(0, total_needed - old_block_size)
    if extra > 0:
        insertion_point = TASK_ROW_END + 1
        ws.insert_rows(insertion_point, amount=extra)
        _shift_merged_ranges_below(ws, insertion_point, extra)
        for row in ws.iter_rows(min_row=insertion_point + extra, max_row=ws.max_row):
            old_row_number = row[0].row - extra
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    cell.value = _reindex_formula_self_refs(cell.value, old_row_number, cell.row)
    elif total_needed < old_block_size:
        warnings.append(
            f"This RFQ only needed {total_needed} rows but the template block reserves {old_block_size} - "
            f"there'll be a few blank spacer rows before the Variation 1 section. Harmless, just a cosmetic gap."
        )

    row_cursor = TASK_ROW_START
    item_counter = 1
    phase_subtotal_rows = []

    for phase in phases:
        header_row = row_cursor
        for col in ["A", DESCRIPTION_COL] + HOURS_COLS + FEE_COLS + [EXPENSE_COL, ROW_SUBTOTAL_COL]:
            ws[f"{col}{header_row}"] = None
        ws[f"{DESCRIPTION_COL}{header_row}"] = phase.get("phase_name", "Phase")
        ws[f"{DESCRIPTION_COL}{header_row}"].font = copy.copy(header_style.font)
        row_cursor += 1

        phase_start = row_cursor
        for description in phase.get("deliverables", []):
            ws[f"A{row_cursor}"] = item_counter
            ws[f"A{row_cursor}"]._style = copy.copy(task_style["A"]._style)
            ws[f"{DESCRIPTION_COL}{row_cursor}"] = description
            ws[f"{DESCRIPTION_COL}{row_cursor}"]._style = copy.copy(task_style[DESCRIPTION_COL]._style)
            for hours_col, fee_col in zip(HOURS_COLS, FEE_COLS):
                ws[f"{fee_col}{row_cursor}"] = f"={hours_col}{row_cursor}*{hours_col}$8"
                ws[f"{fee_col}{row_cursor}"]._style = copy.copy(task_style[fee_col]._style)
                ws[f"{hours_col}{row_cursor}"]._style = copy.copy(task_style[hours_col]._style)
            ws[f"{EXPENSE_COL}{row_cursor}"]._style = copy.copy(task_style[EXPENSE_COL]._style)
            sum_terms = "+".join(f"{c}{row_cursor}" for c in [EXPENSE_COL] + FEE_COLS)
            ws[f"{ROW_SUBTOTAL_COL}{row_cursor}"] = f"=(SUM({sum_terms}))"
            ws[f"{ROW_SUBTOTAL_COL}{row_cursor}"]._style = copy.copy(task_style[ROW_SUBTOTAL_COL]._style)
            item_counter += 1
            row_cursor += 1
        phase_end = row_cursor - 1

        subtotal_row = row_cursor
        ws[f"{DESCRIPTION_COL}{subtotal_row}"] = f"{phase.get('phase_name', 'Phase')} Subtotal"
        ws[f"{DESCRIPTION_COL}{subtotal_row}"]._style = copy.copy(subtotal_style[DESCRIPTION_COL]._style)
        for col in HOURS_COLS + FEE_COLS + [EXPENSE_COL]:
            ws[f"{col}{subtotal_row}"] = f"=SUM({col}{phase_start}:{col}{phase_end})"
            ws[f"{col}{subtotal_row}"]._style = copy.copy(subtotal_style[col]._style)
        ws[f"{ROW_SUBTOTAL_COL}{subtotal_row}"] = f"=SUM({ROW_SUBTOTAL_COL}{phase_start}:{ROW_SUBTOTAL_COL}{phase_end})"
        ws[f"{ROW_SUBTOTAL_COL}{subtotal_row}"]._style = copy.copy(subtotal_style[ROW_SUBTOTAL_COL]._style)
        phase_subtotal_rows.append(subtotal_row)
        row_cursor += 1

    total_row = row_cursor
    ws[f"{DESCRIPTION_COL}{total_row}"] = "Total (Exc. GST)"
    ws[f"{DESCRIPTION_COL}{total_row}"]._style = copy.copy(subtotal_style[DESCRIPTION_COL]._style)
    ws[f"{ROW_SUBTOTAL_COL}{total_row}"] = "=" + "+".join(f"{ROW_SUBTOTAL_COL}{r}" for r in phase_subtotal_rows)
    ws[f"{ROW_SUBTOTAL_COL}{total_row}"]._style = copy.copy(subtotal_style[ROW_SUBTOTAL_COL]._style)

    # Fix the Jessica Ward role mislabel in Variation 1, same as before -
    # search for it since row insertion moved it.
    fixed_mislabel = False
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "Project Engineer" and ws.cell(row=cell.row - 1, column=cell.column).value == "Jessica Ward":
                cell.value = "Hydrologist and Technician"
                fixed_mislabel = True
    if not fixed_mislabel:
        warnings.append("Expected to find and fix a 'Project Engineer' mislabel under Jessica Ward but didn't find it - check the Variation 1 block manually.")

    wb.save(output_path)
    return warnings


def rough_fee_estimate(extracted: dict, hours_per_deliverable_range: tuple[int, int] = (8, 20)) -> str:
    """
    Returns a plain-text rough estimate for the proposal summary / review
    notification - NEVER written into the spreadsheet.

    This is a planning aid, not a quote. It assumes a generic hours-per-
    deliverable range applied at the blended average rate across the whole
    team, because the RFQ extraction deliberately does not guess actual
    staffing/hours - that's a scoping judgment call for the reviewer.
    Adjust hours_per_deliverable_range if 8-20hrs/task doesn't match your
    typical task size. Reports a line per phase plus a grand total.
    """
    phases = extracted.get("phases", [])
    if not phases:
        return "No deliverables extracted - cannot produce a rough estimate."

    blended_rate = sum(STAFF_RATES) / len(STAFF_RATES)
    low_hrs, high_hrs = hours_per_deliverable_range
    lines = ["Rough order-of-magnitude estimate (NOT a quote, NOT in the spreadsheet):"]
    grand_low = grand_high = 0.0
    grand_n = 0

    for phase in phases:
        n = len(phase.get("deliverables", []))
        low_total = n * low_hrs * blended_rate
        high_total = n * high_hrs * blended_rate
        grand_low += low_total
        grand_high += high_total
        grand_n += n
        lines.append(f"  {phase.get('phase_name', 'Phase')}: {n} tasks -> ${low_total:,.0f} - ${high_total:,.0f}")

    lines.append(f"  TOTAL: {grand_n} tasks x {low_hrs}-{high_hrs} hrs/task x ${blended_rate:.0f}/hr blended rate = ${grand_low:,.0f} - ${grand_high:,.0f} (Exc. GST)")
    lines.append(
        f"  Assumption: {low_hrs}-{high_hrs} hrs/task is a generic placeholder, not based on this RFQ's actual "
        f"complexity. Adjust the range in rough_fee_estimate() if it doesn't match typical task sizing."
    )
    return "\n".join(lines)


if __name__ == "__main__":
    sample_extracted = {
        "project_title": "Elizabeth Street Flood Mitigation Concept Design",
        "phases": [
            {
                "phase_name": "Phase 1 - Background and Modelling",
                "deliverables": [
                    "Review supplied TUFLOW/ARR2019 hazard and flow grids",
                    "Identify catchment-wide flood storage opportunity zones",
                ],
            },
            {
                "phase_name": "Phase 2 - Concept and Reporting",
                "deliverables": [
                    "Concept-level cost estimate for prioritised storage sites",
                    "Draft Flood Mitigation Toolkit GIS deliverable",
                    "Stakeholder workshop and comment integration",
                ],
            },
        ],
    }
    warnings = fill_phased_template(
        "Rain_Project_Fee_Tracker_Template.xlsx",
        "Rain_Project_Fee_Tracker_FILLED.xlsx",
        sample_extracted,
    )
    for w in warnings:
        print("WARNING:", w)
    print()
    print(rough_fee_estimate(sample_extracted))
